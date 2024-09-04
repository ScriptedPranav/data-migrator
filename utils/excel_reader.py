import openpyxl

def read_excel(file_path, sheet_name, column_mapping):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    
    data = []
    headers = {v: i for i, v in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))}
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        is_empty_row = True
        
        for excel_col, json_key in column_mapping.items():
            if excel_col in headers:
                cell_value = row[headers[excel_col]]
                if cell_value is not None and cell_value != "":
                    is_empty_row = False
                row_data[json_key] = cell_value
        
        if not is_empty_row:
            data.append(row_data)
    
    return data

